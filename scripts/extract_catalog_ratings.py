"""
Part 1 — Extract Altivar Process ATV600 catalog rating tables to CSV/XLSX.

Source: Catalog Altivar Process ATV600 variable speed drives.pdf
Primary KPI: drive commercial Reference (ATV630 / ATV650 / ATV680).
Secondary: passive filter accessories (VW3A461…).
DC choke: feature flag only (no DC-choke accessory SKU table in client PDFs).
"""

from __future__ import annotations

import csv
import re
from pathlib import Path
from typing import Any

import fitz
from openpyxl import Workbook
from openpyxl.styles import Font

ROOT = Path(__file__).resolve().parent.parent
CATALOG = ROOT / "Catalog Altivar Process ATV600 variable speed drives.pdf"
OUT_DIR = ROOT / "data" / "catalog"

# 0-based page indices: wall/floor/cabinet + ATV680 LH + ATV660 Compact
DRIVE_PAGES = list(range(23, 31)) + [85, 86] + [112, 113] + [118, 119]  # pp.24-31, 86-87, 113-114, 119-120
FILTER_PAGES = [51, 52, 53, 54]  # PDF pp.52-55

REF_RE = re.compile(r"^(ATV(?:630|650|660|680)[A-Z0-9]+)")
FILTER_RE = re.compile(r"^(VW3A461\d{2})")
NUM_RE = re.compile(r"^[\d]+(?:[.,]\d+)?$")
DASHES = {"–", "-", "—", "−", ""}


def _norm_num(tok: str) -> float | None:
    tok = tok.strip().replace(",", ".")
    if tok in DASHES or tok == "":
        return None
    try:
        return float(tok)
    except ValueError:
        return None


def _lines(page: fitz.Page) -> list[str]:
    text = page.get_text("text")
    # soft hyphen / weird spaces
    text = text.replace("\u00ad", "").replace("\xa0", " ")
    out: list[str] = []
    for raw in text.splitlines():
        s = raw.strip()
        if not s:
            continue
        # strip footnote markers glued to refs: ATV630C11N4 (6)
        s = re.sub(r"\s*\(\d+\)\s*$", "", s).strip()
        out.append(s)
    return out


def _page_meta(lines: list[str], page_1based: int) -> dict[str, str]:
    blob = " ".join(lines[:40]).lower()
    supply = ""
    m = re.search(
        r"(200\.\.\.240|380\.\.\.480|380\.\.\.440|380\.\.\.415|500\.\.\.690|480)\s*v",
        blob,
        re.I,
    )
    if m:
        supply = m.group(1).replace("...", "-") + " V"
    elif "480 v" in blob:
        supply = "480 V"

    if "cabinet integration" in blob or "for integration" in blob:
        mounting = "cabinet_Z"
    elif "floor-standing" in blob:
        mounting = "floor"
    elif "low harmonic" in blob:
        mounting = "system_LH"
    elif "compact drive" in blob:
        mounting = "system_compact"
    else:
        mounting = "wall"

    ip_note = ""
    for key in ("IP21", "IP55", "IP54", "IP20", "IP23", "IP00", "UL Type 1"):
        if key.lower() in blob:
            ip_note = key
            break

    emc_note = ""
    if "category c2" in blob and "category c3" in blob:
        emc_note = "C2 or C3"
    elif "category c2" in blob:
        emc_note = "C2"
    elif "category c3" in blob:
        emc_note = "C3"

    return {
        "supply_v": supply,
        "mounting": mounting,
        "ip_note": ip_note,
        "emc_note": emc_note,
        "source_page": str(page_1based),
    }


def _family_from_ref(ref: str) -> str:
    if ref.startswith("ATV680"):
        return "ATV680"
    if ref.startswith("ATV660"):
        return "ATV660"
    if ref.startswith("ATV650"):
        return "ATV650"
    return "ATV630"


def _has_dc_choke(ref: str, mounting: str, supply_v: str) -> str:
    """
    Feature flag — not a SKU.
    Catalog p.17: embedded DC choke for 380…480 V Process offer.
    ATV680 LH systems use different harmonic mitigation → no.
    ATV660 compact systems include line reactor (not classic DC choke SKU).
    """
    if ref.startswith("ATV680"):
        return "false"
    if ref.startswith("ATV660"):
        return "false"
    if "500-690" in supply_v or "Y6" in ref:
        return "unknown"
    if ref.startswith("ATV630") or ref.startswith("ATV650"):
        return "true"
    return "unknown"


def _default_variant_rank(ref: str, mounting: str) -> int:
    """Lower = preferred default (normal Process before cabinet Z)."""
    if mounting == "cabinet_Z" or ref.endswith("Z"):
        return 30
    if ref.endswith("E"):
        return 20
    if ref.endswith("F"):
        return 15
    if ref.startswith("ATV650"):
        return 10
    if ref.startswith("ATV680"):
        return 5
    return 0  # standard ATV630 wall/floor


def extract_drive_rows(doc: fitz.Document) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for p0 in DRIVE_PAGES:
        page = doc[p0]
        lines = _lines(page)
        meta = _page_meta(lines, p0 + 1)
        # Skip 500-690 V dual-column page in v1 (different layout) — still try simple pairs
        i = 0
        pending_nd: dict[str, Any] | None = None
        while i < len(lines):
            if lines[i] not in ("ND", "HD"):
                i += 1
                continue
            duty = lines[i]
            i += 1
            nums: list[str] = []
            while i < len(lines) and (
                NUM_RE.match(lines[i].replace(",", "")) or lines[i] in DASHES
            ):
                # stop if this looks like a weight fragment alone later
                nums.append(lines[i])
                i += 1
            ref = None
            if i < len(lines):
                m = REF_RE.match(lines[i].replace(" ", ""))
                if m:
                    ref = m.group(1)
                    i += 1
                    # optional weight lines like 4.500/ then 9.921
                    while i < len(lines) and (
                        "/" in lines[i] or NUM_RE.match(lines[i].replace(",", ""))
                    ):
                        # don't consume next ND/HD
                        if lines[i] in ("ND", "HD"):
                            break
                        if REF_RE.match(lines[i].replace(" ", "")):
                            break
                        if FILTER_RE.match(lines[i]):
                            break
                        i += 1
                        # only skip a couple weight tokens
                        if i < len(lines) and lines[i] in ("ND", "HD"):
                            break
                        # limit runaway
                        if len(nums) > 20:
                            break

            # Parse kW/HP from start of nums (standard tables)
            # 500-690 V tables have 4 leading power values — mark and take first pair for ND at 500V
            kw = _norm_num(nums[0]) if len(nums) >= 1 else None
            hp = _norm_num(nums[1]) if len(nums) >= 2 else None
            # Heuristic: if page is 500-690, nums[0:4] are dual voltage powers
            dual = "500-690" in meta["supply_v"]
            if dual and len(nums) >= 4:
                # keep 500 V pair as primary kw/hp; store 690 note in extra via hp already
                pass

            line_a = _norm_num(nums[2]) if len(nums) >= 3 else None
            line_b = _norm_num(nums[3]) if len(nums) >= 4 else None
            # ATV680 / ATV660 system tables: ND kw, line, kVA, Isc, cont, transient (no HP)
            is_system = meta["mounting"] in ("system_LH", "system_compact") or (
                ref and (ref.startswith("ATV680") or ref.startswith("ATV660"))
            )
            if is_system or (ref and (ref.startswith("ATV680") or ref.startswith("ATV660"))):
                # nums: kw, line, kva, isc, cont, transient — no HP
                kw = _norm_num(nums[0]) if nums else None
                hp = None
                line_a = _norm_num(nums[1]) if len(nums) >= 2 else None
                line_b = None
                cont = _norm_num(nums[4]) if len(nums) >= 5 else None
                trans = _norm_num(nums[5]) if len(nums) >= 6 else None
            elif dual:
                cont = _norm_num(nums[8]) if len(nums) >= 9 else None
                trans = _norm_num(nums[9]) if len(nums) >= 10 else None
                line_a = _norm_num(nums[4]) if len(nums) >= 5 else None
                line_b = _norm_num(nums[5]) if len(nums) >= 6 else None
            else:
                # standard: kw hp line380 line480 kva isc cont transient
                cont = _norm_num(nums[6]) if len(nums) >= 7 else None
                trans = _norm_num(nums[7]) if len(nums) >= 8 else None

            if duty == "ND" and ref:
                mounting = meta["mounting"]
                if ref.endswith("Z"):
                    mounting = "cabinet_Z"
                elif ref.endswith("F"):
                    mounting = "floor"
                elif ref.startswith("ATV660"):
                    mounting = "system_compact"
                elif ref.startswith("ATV680"):
                    mounting = "system_LH"
                row = {
                    "family": _family_from_ref(ref),
                    "supply_v": meta["supply_v"],
                    "mounting": mounting,
                    "duty": "ND",
                    "kw": kw,
                    "hp": hp,
                    "reference": ref,
                    "line_current_a": line_a,
                    "line_current_b": line_b,
                    "max_continuous_a": cont,
                    "max_transient_60s_a": trans,
                    "has_dc_choke": _has_dc_choke(ref, mounting, meta["supply_v"]),
                    "emc_note": meta["emc_note"],
                    "ip_note": meta["ip_note"],
                    "variant_rank": _default_variant_rank(ref, mounting),
                    "is_default_variant": "",  # filled later
                    "source_pdf": CATALOG.name,
                    "source_page": meta["source_page"],
                }
                rows.append(row)
                pending_nd = row
                # also prepare HD sibling when we see HD
            elif duty == "HD":
                # HD may or may not have its own ref (usually shares ND ref)
                use_ref = ref or (pending_nd["reference"] if pending_nd else None)
                if not use_ref or kw is None:
                    continue
                mounting = meta["mounting"]
                if use_ref.endswith("Z"):
                    mounting = "cabinet_Z"
                elif use_ref.endswith("F"):
                    mounting = "floor"
                elif use_ref.startswith("ATV660"):
                    mounting = "system_compact"
                elif use_ref.startswith("ATV680"):
                    mounting = "system_LH"
                if is_system or use_ref.startswith("ATV680") or use_ref.startswith("ATV660"):
                    kw = _norm_num(nums[0]) if nums else None
                    hp = None
                    line_a = _norm_num(nums[1]) if len(nums) >= 2 else None
                    line_b = None
                    cont = _norm_num(nums[4]) if len(nums) >= 5 else None
                    trans = _norm_num(nums[5]) if len(nums) >= 6 else None
                rows.append(
                    {
                        "family": _family_from_ref(use_ref),
                        "supply_v": meta["supply_v"],
                        "mounting": mounting,
                        "duty": "HD",
                        "kw": kw,
                        "hp": hp,
                        "reference": use_ref,
                        "line_current_a": line_a,
                        "line_current_b": line_b,
                        "max_continuous_a": cont,
                        "max_transient_60s_a": trans,
                        "has_dc_choke": _has_dc_choke(
                            use_ref, mounting, meta["supply_v"]
                        ),
                        "emc_note": meta["emc_note"],
                        "ip_note": meta["ip_note"],
                        "variant_rank": _default_variant_rank(use_ref, mounting),
                        "is_default_variant": "",
                        "source_pdf": CATALOG.name,
                        "source_page": meta["source_page"],
                    }
                )
        # end while
    return _mark_default_variants(_dedupe(rows))


def _dedupe(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    seen: set[tuple] = set()
    out = []
    for r in rows:
        key = (r["reference"], r["duty"], r["kw"], r["source_page"])
        if key in seen:
            continue
        seen.add(key)
        out.append(r)
    return out


def _mark_default_variants(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """
    For each (duty, kw, supply bucket, family group), mark lowest variant_rank
    as default. Cabinet Z is never default unless it is the only option.
    """
    # Group by duty + kw (rounded) + rough supply class
    groups: dict[tuple, list[dict[str, Any]]] = {}
    for r in rows:
        if r["kw"] is None:
            r["is_default_variant"] = "false"
            continue
        supply_key = r["supply_v"] or ""
        key = (r["duty"], round(float(r["kw"]), 3), supply_key)
        groups.setdefault(key, []).append(r)

    for group in groups.values():
        # Prefer non-cabinet
        ranked = sorted(group, key=lambda x: (x["variant_rank"], x["reference"]))
        best = ranked[0]["reference"]
        for r in group:
            r["is_default_variant"] = "true" if r["reference"] == best else "false"
    return rows


def extract_passive_filters(doc: fitz.Document) -> list[dict[str, Any]]:
    """
    Passive filter tables: motor kW/HP may share one VW3 filter across several drives.
    We emit one row per (motor_kw, drive_ref, filter_ref) association when possible,
    or one row per filter block with for_drive_refs joined.
    """
    rows: list[dict[str, Any]] = []
    for p0 in FILTER_PAGES:
        page = doc[p0]
        lines = _lines(page)
        blob_head = " ".join(lines[:30]).lower()
        supply = "400 V 50 Hz" if "400 v" in blob_head else ""
        if "460 v" in blob_head:
            supply = "460 V 60 Hz"
        thdi = ""
        if "thdi < 5%" in blob_head or "thdi <5%" in blob_head:
            thdi = "<5%"
        elif "thdi < 10%" in blob_head:
            thdi = "<10%"

        # Walk: collect pending motor+drives until VW3 appears
        i = 0
        pending: list[dict[str, Any]] = []
        while i < len(lines):
            # motor kW line start: number then maybe HP then ATV refs
            if NUM_RE.match(lines[i].replace(",", "")) and i + 1 < len(lines):
                kw = _norm_num(lines[i])
                hp = None
                j = i + 1
                if j < len(lines) and (
                    NUM_RE.match(lines[j].replace(",", "")) or lines[j] in DASHES
                ):
                    hp = _norm_num(lines[j])
                    j += 1
                drives: list[str] = []
                while j < len(lines):
                    m = REF_RE.match(lines[j].replace(" ", ""))
                    if m:
                        drives.append(m.group(1))
                        j += 1
                        continue
                    break
                if drives:
                    pending.append({"kw": kw, "hp": hp, "drives": drives})
                    i = j
                    continue

            m = FILTER_RE.match(lines[i])
            if m:
                filt = m.group(1)
                # look back for qty and currents: typically ... A A qty VW3 weight
                # scan previous non-ATV numeric tokens
                qty = 1
                in_a = None
                out_a = None
                # previous lines may be weight starting after — currents before filter
                back = i - 1
                tokens: list[str] = []
                while back >= 0 and len(tokens) < 6:
                    if REF_RE.match(lines[back].replace(" ", "")):
                        break
                    if lines[back] in ("ND", "HD"):
                        break
                    tokens.insert(0, lines[back])
                    back -= 1
                # tokens often: in_a out_a qty   OR just qty
                nums = []
                for t in tokens:
                    if NUM_RE.match(t.replace(",", "")):
                        nums.append(_norm_num(t))
                if len(nums) >= 3:
                    in_a, out_a, qty = nums[-3], nums[-2], int(nums[-1] or 1)
                elif len(nums) == 1:
                    qty = int(nums[0] or 1)

                if not pending:
                    rows.append(
                        {
                            "motor_kw": None,
                            "motor_hp": None,
                            "for_drive_refs": "",
                            "filter_current_in_a": in_a,
                            "filter_current_out_a": out_a,
                            "quantity_per_drive": qty,
                            "filter_reference": filt,
                            "supply": supply,
                            "thdi_note": thdi,
                            "source_pdf": CATALOG.name,
                            "source_page": str(p0 + 1),
                        }
                    )
                else:
                    for block in pending:
                        rows.append(
                            {
                                "motor_kw": block["kw"],
                                "motor_hp": block["hp"],
                                "for_drive_refs": "|".join(block["drives"]),
                                "filter_current_in_a": in_a,
                                "filter_current_out_a": out_a,
                                "quantity_per_drive": qty,
                                "filter_reference": filt,
                                "supply": supply,
                                "thdi_note": thdi,
                                "source_pdf": CATALOG.name,
                                "source_page": str(p0 + 1),
                            }
                        )
                    pending = []
                i += 1
                continue
            i += 1
    return rows


def _write_csv(path: Path, rows: list[dict[str, Any]], fieldnames: list[str]) -> None:
    tmp = path.with_suffix(path.suffix + ".tmp")
    with tmp.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
        w.writeheader()
        for r in rows:
            w.writerow(r)
    try:
        tmp.replace(path)
    except PermissionError:
        alt = path.with_name(path.stem + "_new" + path.suffix)
        tmp.replace(alt)
        print(f"WARNING: {path.name} locked; wrote {alt.name} instead")


def _write_xlsx(
    path: Path,
    drive_rows: list[dict[str, Any]],
    filter_rows: list[dict[str, Any]],
    drive_fields: list[str],
    filter_fields: list[str],
) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "drive_ratings"
    ws.append(drive_fields)
    for c in ws[1]:
        c.font = Font(bold=True)
    for r in drive_rows:
        ws.append([r.get(k) for k in drive_fields])

    ws2 = wb.create_sheet("passive_filters")
    ws2.append(filter_fields)
    for c in ws2[1]:
        c.font = Font(bold=True)
    for r in filter_rows:
        ws2.append([r.get(k) for k in filter_fields])

    notes = wb.create_sheet("NOTES")
    notes["A1"] = "Part 1 catalog extract notes"
    notes["A1"].font = Font(bold=True)
    notes_lines = [
        "Primary KPI = drive commercial Reference.",
        "Variant default: is_default_variant=true prefers normal Process (not …Z cabinet) for same kW/duty/supply.",
        "DC choke: has_dc_choke is a FEATURE FLAG only. Client PDFs have no DC-choke accessory SKU table.",
        "Catalog p.17: embedded DC choke on 380…480 V Process offer; ATV680/ATV660 marked false.",
        "Passive filters: secondary accessory KPI (VW3A461…).",
        "Getting Started PDF: not used (no rating tables).",
        "Installation manual: reserved for Part 2 gold validation cross-check.",
        f"Source: {CATALOG.name}",
        "Drive pages extracted: 24-31, 86-87, 113-114, 119-120",
        "Filter pages extracted: 52-55",
    ]
    for i, line in enumerate(notes_lines, start=3):
        notes[f"A{i}"] = line
    wb.save(path)


def main() -> None:
    if not CATALOG.exists():
        raise SystemExit(f"Catalog PDF not found: {CATALOG}")
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    doc = fitz.open(CATALOG)

    drive_rows = extract_drive_rows(doc)
    filter_rows = extract_passive_filters(doc)

    drive_fields = [
        "family",
        "supply_v",
        "mounting",
        "duty",
        "kw",
        "hp",
        "reference",
        "line_current_a",
        "line_current_b",
        "max_continuous_a",
        "max_transient_60s_a",
        "has_dc_choke",
        "emc_note",
        "ip_note",
        "variant_rank",
        "is_default_variant",
        "source_pdf",
        "source_page",
    ]
    filter_fields = [
        "motor_kw",
        "motor_hp",
        "for_drive_refs",
        "filter_current_in_a",
        "filter_current_out_a",
        "quantity_per_drive",
        "filter_reference",
        "supply",
        "thdi_note",
        "source_pdf",
        "source_page",
    ]

    _write_csv(OUT_DIR / "drive_ratings.csv", drive_rows, drive_fields)
    _write_csv(OUT_DIR / "passive_filters.csv", filter_rows, filter_fields)
    _write_xlsx(
        OUT_DIR / "catalog_kpis.xlsx",
        drive_rows,
        filter_rows,
        drive_fields,
        filter_fields,
    )

    notes = OUT_DIR / "EXTRACT_NOTES.md"
    notes.write_text(
        "\n".join(
            [
                "# Catalog KPI extract — Part 1",
                "",
                f"- Source: `{CATALOG.name}`",
                f"- Drive rating rows: **{len(drive_rows)}**",
                f"- Passive filter rows: **{len(filter_rows)}**",
                "- Outputs: `drive_ratings.csv`, `passive_filters.csv`, `catalog_kpis.xlsx`",
                "",
                "## Scope",
                "- Drive tables: PDF pages 24–31, 86–87, 113–114 (ATV660), 119–120 (ATV680)",
                "- Passive filters: PDF pages 52–55",
                "",
                "## DC choke",
                "- **No DC-choke accessory part-number table** in the three client PDFs.",
                "- Column `has_dc_choke` is a **feature flag** (true for typical ATV630/650 Process 200–480 V; false for ATV680 LH / ATV660; unknown for 500–690 V / Y6).",
                "",
                "## Variant default (approved)",
                "- `is_default_variant=true` prefers normal Process refs over `…Z` cabinet for the same kW/duty/supply.",
                "- Cabinet codes stay in the dataset for when the user asks for cabinet integration.",
                "",
                "## ATV660",
                "- Compact Drive Systems tables on Catalog pp.113–114.",
                "- Used for Standard harmonics >315 kW (and available in 110–315 band; sizing prefers ATV630 there).",
                "",
                "## Next",
                "- Part 2: gold-validate ~20–30 rows against the PDF by eye.",
                "",
            ]
        ),
        encoding="utf-8",
    )

    # Quick sanity prints
    sample = [r for r in drive_rows if r["reference"] == "ATV630D22N4" and r["duty"] == "ND"]
    atv660 = [r for r in drive_rows if (r.get("family") or "").startswith("ATV660")]
    print(f"drive_rows={len(drive_rows)} filter_rows={len(filter_rows)} atv660_rows={len(atv660)}")
    print("sample ATV630D22N4 ND:", sample[:1])
    print("sample ATV660:", atv660[:2])
    defaults_22 = [
        r
        for r in drive_rows
        if r.get("kw") == 22.0 and r["duty"] == "ND" and "380" in (r["supply_v"] or "")
    ]
    print("22 kW ND 380V variants:", [(r["reference"], r["is_default_variant"]) for r in defaults_22])


if __name__ == "__main__":
    main()
